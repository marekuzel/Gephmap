from openpyxl import Workbook, load_workbook

#cycle for counting number of appearances for Groups, in excel this is done by COUNTIF
def countGroupAppearances(ws, group_names, group_dict_appearances):
    for cell in ws["A:A"]:
        if cell.value != "Page Name":
            if cell.value not in group_names:
                group_names.append(cell.value)
            if cell.value not in group_dict_appearances:
                group_dict_appearances[cell.value] = 1
            else:
                group_dict_appearances[cell.value] += 1

def countGroupInteractions(ws, group_dict_appearances, group_dict_totInt):
    for node in group_dict_appearances:
        for i, cell in enumerate(ws["A:A"]):
            i += 1
            if node not in group_dict_totInt:
                group_dict_totInt[node] = 0
            if (node == cell.value and i>1):#i must be highere then 1 because openpyxl cant handle cell value of 0
                group_dict_totInt[node] += int(ws.cell(row=i, column=14).value) #column N always contains number of interaction for inidividual node




def countLinkAppearances(ws, link_names, link_dict_appearances):
    for cell in ws["AF:AF"]:
        if cell.value != "Link" or cell.value != "None":
            if cell.value not in link_names:
                link_names.append(cell.value)
        if cell.value not in link_dict_appearances:
            link_dict_appearances[cell.value] = 1
        else:
            link_dict_appearances[cell.value] += 1


def countLinkInteractions(ws, link_dict_appearances, link_dict_totInt):
    for link in link_dict_appearances:
        for i, cell in enumerate(ws["AF:AF"]):
            i += 1
            if link not in link_dict_totInt:
                link_dict_totInt[link] = 0
            if (link == cell.value and i>1):#i must be highere then 1 because openpyxl cant handle cell value of 0
                link_dict_totInt[link] += int(ws.cell(row=i, column=14).value) #column N always contains number of interaction for inidividual link

def CSV_countGroupAppearances(csv, group_names, group_dict_appearances):
    with open (csv) as csvfile:
        for row in csvfile:
            if row[0] != "Page Name":
                if row[0] not in group_names:
                    group_names.append(row[0])
                if row[0] not in group_dict_appearances:
                    group_dict_appearances[row[0]] = 1
                else:
                    group_dict_appearances[row[0]] += 1

def CSV_countGroupInteractions(csv, group_dict_appearances, group_dict_totInt):
    for node in group_dict_appearances:
        with open (csv) as csvfile:
            for row in csvfile:
                if node not in group_dict_totInt:
                    group_dict_totInt[node] = 0
                if (node == row[0] and row[0] != "Page Name"):
                    group_dict_totInt[node] += int(row[13])
       

def CSV_countLinkAppearances(csv, group_dict_appearances):
    with open (csv) as csvfile:
        for  row in csvfile:
            if row[31] != "Link" or row[31] != "None":
                if row[31] not in group_dict_appearances:
                    group_dict_appearances[row[31]] = 1
                else:
                    group_dict_appearances[row[31]] += 1

def CSV_countLinkInteractions(csv, group_dict_appearances, group_dict_totInteractions):
    for link in group_dict_appearances:
        with open (csv) as csvfile:
            for row in csvfile:
                if link not in group_dict_totInteractions:
                    group_dict_totInteractions[link] = 0
                if (link == row[31] and row[31] != "Link"):
                    group_dict_totInteractions[link] += int(row[13])