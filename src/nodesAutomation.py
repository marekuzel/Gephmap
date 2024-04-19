from openpyxl import Workbook, load_workbook
from excelCounter import *


def main ():
    file = returnFile()

    print ("Starting " + file)
    wb = load_workbook(file)
    ws = wb.active
    #group_dict_appearances is used to store amount of unique nodes
    #group_dict_totInt stores the number of interactions each unique node recieved
    group_names = []
    group_dict_appearances = {}
    group_dict_totInt = {}
    link_names = []
    link_dict_appearances = {}
    link_dict_totInt = {}  

    countGroupAppearances(ws,group_names, group_dict_appearances)
    print ("Group Appearances done in " + file)
    countGroupInteractions(ws, group_dict_appearances, group_dict_totInt)
    print ("Group Interactions done" + file)
    countLinkAppearances(ws, link_names, link_dict_appearances)
    print ("Link Appearances done" + file)
    countLinkInteractions(ws, link_dict_appearances, link_dict_totInt)
    print ("Link Interactions done" + file)

    with open(file + "_Nodes.csv", 'w', encoding="utf-8") as f:
        f.write ("Label, Id, Appearances, Total Interactions, Type\n")
        for row in group_names:
            # Write the values in a row.
            f.write('"{}"'.format(str(row).replace('"', '')))
            f.write(",")
            f.write('"{}"'.format(str(row).replace('"', '')))
            f.write(",")
            f.write(str(group_dict_appearances[row]))
            f.write(",")
            f.write(str(group_dict_totInt[row]))
            f.write(',Group\n') # Add a new line
        for row in link_names:
            f.write('"{}"'.format(str(row).replace('"', '')))
            f.write(",")
            f.write('"{}"'.format(str(row).replace('"', '')))
            f.write(",")
            f.write(str(link_dict_appearances[row]))
            f.write(",")
            f.write(str(link_dict_totInt[row]))
            f.write(',Link\n') # Add a new line

    with open(file + "_Edges.csv", 'w', encoding="utf-8") as f:
        f.write ("Source, Target\n")
        i=2
        for cell in ws["A:A"]:
            f.write('"{}"'.format(str(ws.cell(row=i, column=1).value).replace('"', '')))
            f.write(",")
            f.write('"{}"'.format(str(ws.cell(row=i, column=32).value).replace('"', '')))
            i+=1
            f.write("\n")